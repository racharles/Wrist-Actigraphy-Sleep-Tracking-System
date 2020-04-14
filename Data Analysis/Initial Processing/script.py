# Data Processing
#
# Reads and processes data from an excel worksheet
# Writes median, mean, max, 95% of each 3-min epoch to another excel worksheet
#
# Instructions:
# in the main program at the bottom, fill in names and tracking_num with the desired names.
# An excel worksheet with the name "processed_[tracking_num].xlsx" will appear
# in the current working directory.

import math
import numpy as np
import openpyxl
import matplotlib.pyplot as plt


def read_data(sheet):
    """Read data from excel sheet and return dictionary with data arranged in 3-min epochs"""

    data = {}
    count = 0
    epoch_num = 1
    time, accX, accY, accZ, act = ([] for i in range(5))

    for i in range(2,sheet.max_row):
        # check if the data should go into the next epoch
        if count > 2:
            # put data values into each epoch
            data[epoch_num] = {"time": time, "accX": accX,
                                "accY": accY, "accZ": accZ, "act": act}
            count = 0
            time, accX, accY, accZ, act = ([] for i in range(5))
            epoch_num += 1
        else:
            count += 1  # update line counter

        # read out the x, y, z values
        time.append(sheet.cell(i,1).value)
        accX.append(sheet.cell(i,2).value)
        accY.append(sheet.cell(i,3).value)
        accZ.append(sheet.cell(i,4).value)

        # compose x(t), y(t), z(t) into act(t)
        act.append(math.sqrt(accX[-1]**2 + accY[-1]**2 + accZ[-1]**2))

    data[epoch_num] = {"time": time, "accX": accX,
                         "accY": accY, "accZ": accZ, "act": act}
    return data


def process_data(data):
    """Processes data from read_data to find median, mean, max, and 95% of the data"""

    processed_data = {}

    for i in range(len(data)):
        # find median, mean, max value from each epoch
        median = np.median(data[i+1]["act"])
        mean = np.mean(data[i+1]["act"])
        max_value = max(data[i+1]["act"])
        percent_95 = np.percentile(data[i+1]["act"], 95)

        processed_data[i+1] = {"median":median, "mean":mean, "max":max_value, "95%":percent_95}
    return processed_data


def write_data(sheet, tracking_num):
    # create a new workbook and open worksheet
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # headers for data
    new_sheet.cell(1, 1).value = "Tracking_num"
    new_sheet.cell(1, 2).value = "Epoch"
    new_sheet.cell(1, 3).value = "Median"
    new_sheet.cell(1, 4).value = "Mean"
    new_sheet.cell(1, 5).value = "Max"
    new_sheet.cell(1, 6).value = "95%"

    raw_data = read_data(sheet)
    processed_data = process_data(read_data(sheet))

    # write data into the worksheet
    for i in range(2, len(processed_data)+2):
        new_sheet.cell(i, 1).value = tracking_num
        new_sheet.cell(i, 2).value = i-1
        new_sheet.cell(i, 3).value = processed_data[i-1]["median"]
        new_sheet.cell(i, 4).value = processed_data[i-1]["mean"]
        new_sheet.cell(i, 5).value = processed_data[i-1]["max"]
        new_sheet.cell(i, 6).value = processed_data[i-1]["95%"]

    # save worksheet to working directory
    new_wb.save("processed_" + tracking_num + ".xlsx")


if __name__ == "__main__":
    # insert names of the data files
    # person_trial = ["p1n1", "p1n2", "p1n3", "p2n1", "p2n2", "p2n3",
    #                 "p3n1", "p3n2", "p5n1", "p5n3", "p6n1", "p6n2", "p7n1", "p7n2"]
    person_trial = ["p7n1", "p7n2"]

    # wb = openpyxl.load_workbook("p6n2.xlsx")
    # sheet = wb.active

    # data = process_data(read_data(sheet))

    for name in person_trial:
        print("Currently processing p" + name[1] + "n" + name[3] + "...")

        # open workbook with data
        wb = openpyxl.load_workbook(name + ".xlsx")
        sheet = wb.active

        data = process_data(read_data(sheet))

        # read and process data
        write_data(sheet, "p" + name[1] + "n" + name[3])
        print("Finished p" + name[1] + "n" + name[3])
